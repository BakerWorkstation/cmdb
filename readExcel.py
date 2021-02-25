#!/usr/bin/env python
# -*- coding:UTF-8 -*-

import os
from openpyxl import load_workbook


def patchData(filename):
    # 打开一个workbook
    
    wb = load_workbook(filename=filename)
    
    # 获取当前活跃的worksheet,默认就是第一个worksheet
    # ws = wb.active
    
    # 当然也可以使用下面的方法
    
    # 获取所有表格(worksheet)的名字
    sheets = wb.get_sheet_names()
    # 第一个表格的名称
    sheet_first = sheets[0]
    
    # 获取特定的worksheet
    ws = wb.get_sheet_by_name(sheet_first)
    
    # 获取表格所有行和列，两者都是可迭代的
    rows = ws.rows
    columns = ws.columns
    
    # 迭代所有的行
    result = []
    flag = 1
    for row in rows:
        if flag == 1:
            flag += 1
            continue
        line = [col.value for col in row]
        if  not line[0]:
            flag += 1
            continue
        #for eachcol in line:
        #    print eachcol
        result.append(line)
        flag += 1
    return result
    
    # 通过坐标读取值
    #print ws.cell('A1').value  # A表示列,1表示行
    
    #for i in rows:
    #    for j in columns:
    #        print ws.cell(row=i, column=j).value
    #        pass

#if __name__ == '__main__':
#   patchData('1.xlsx')
